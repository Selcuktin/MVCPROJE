from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import json
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from .models import Course, CourseGroup, Enrollment, Assignment, Submission, Announcement
from students.models import Student
from teachers.models import Teacher
from .forms import CourseForm, CourseGroupForm, AssignmentForm, SubmissionForm, AnnouncementForm, EnrollmentForm, GradeForm

# Course Views
class CourseListView(LoginRequiredMixin, ListView):
    model = Course
    template_name = 'courses/list.html'
    context_object_name = 'courses'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Course.objects.filter(status='active')
        search = self.request.GET.get('search')
        department = self.request.GET.get('department')
        semester = self.request.GET.get('semester')
        
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(department__icontains=search)
            )
        
        if department:
            queryset = queryset.filter(department__icontains=department)
            
        if semester:
            queryset = queryset.filter(semester=semester)
            
        return queryset.order_by('code')

class CourseDetailView(LoginRequiredMixin, DetailView):
    model = Course
    template_name = 'courses/detail.html'
    context_object_name = 'course'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course = self.get_object()
        context['groups'] = course.groups.filter(status='active').select_related('teacher')
        context['can_enroll'] = False
        
        if hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student':
            try:
                student = Student.objects.get(user=self.request.user)
                context['can_enroll'] = True
                context['student'] = student
            except Student.DoesNotExist:
                pass
        
        return context

class CourseCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Course
    form_class = CourseForm
    template_name = 'courses/form.html'
    success_url = reverse_lazy('courses:list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin', 'teacher']
    
    def form_valid(self, form):
        messages.success(self.request, 'Ders başarıyla oluşturuldu.')
        return super().form_valid(form)

class CourseUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Course
    form_class = CourseForm
    template_name = 'courses/form.html'
    success_url = reverse_lazy('courses:list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin', 'teacher']

class CourseDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Course
    template_name = 'courses/delete.html'
    success_url = reverse_lazy('courses:list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin']

# Course Group Views
class CourseGroupListView(LoginRequiredMixin, ListView):
    model = CourseGroup
    template_name = 'courses/group_list.html'
    context_object_name = 'groups'
    paginate_by = 20
    
    def get_queryset(self):
        return CourseGroup.objects.select_related('course', 'teacher').filter(status='active')

class CourseGroupDetailView(LoginRequiredMixin, DetailView):
    model = CourseGroup
    template_name = 'courses/group_detail.html'
    context_object_name = 'group'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.get_object()
        
        # Enrollments ile birlikte notları da getir
        enrollments = group.enrollments.select_related('student')
        
        # Her enrollment için notları getir
        from notes.models import Note
        enrollment_data = []
        for enrollment in enrollments:
            notes = Note.objects.filter(
                student=enrollment.student.user,  # Student nesnesinden User nesnesine geçiş
                course=group.course
            )
            
            # Notları sınav türüne göre grupla
            note_dict = {}
            for note in notes:
                note_dict[note.exam_type] = note
            
            enrollment_data.append({
                'enrollment': enrollment,
                'notes': note_dict
            })
        
        context['enrollment_data'] = enrollment_data
        context['enrollments'] = enrollments  # Geriye uyumluluk için
        context['assignments'] = group.assignments.filter(status='active')
        context['announcements'] = group.announcements.filter(status='active').order_by('-create_date')
        return context

class CourseGroupCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = CourseGroup
    form_class = CourseGroupForm
    template_name = 'courses/group_form.html'
    success_url = reverse_lazy('courses:group_list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin', 'teacher']

class CourseGroupUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CourseGroup
    form_class = CourseGroupForm
    template_name = 'courses/group_form.html'
    success_url = reverse_lazy('courses:group_list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin', 'teacher']

# Enrollment Views
class EnrollmentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Enrollment
    form_class = EnrollmentForm
    template_name = 'courses/enroll.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['group'] = get_object_or_404(CourseGroup, pk=self.kwargs['group_id'])
        return context
    
    def form_valid(self, form):
        group = get_object_or_404(CourseGroup, pk=self.kwargs['group_id'])
        student = Student.objects.get(user=self.request.user)
        
        # Check if already enrolled
        if Enrollment.objects.filter(student=student, group=group).exists():
            messages.error(self.request, 'Bu derse zaten kayıtlısınız.')
            return redirect('courses:detail', pk=group.course.pk)
        
        # Check capacity
        if group.enrollments.count() >= group.course.capacity:
            messages.error(self.request, 'Bu ders grubu dolu.')
            return redirect('courses:detail', pk=group.course.pk)
        
        enrollment = form.save(commit=False)
        enrollment.student = student
        enrollment.group = group
        enrollment.save()
        
        messages.success(self.request, f'{group.course.name} dersine başarıyla kayıt oldunuz.')
        return redirect('courses:detail', pk=group.course.pk)

class EnrollmentListView(LoginRequiredMixin, ListView):
    model = Enrollment
    template_name = 'courses/enrollment_list.html'
    context_object_name = 'enrollments'
    paginate_by = 20
    
    def get_queryset(self):
        if hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student':
            student = Student.objects.get(user=self.request.user)
            return student.enrollments.select_related('group__course', 'group__teacher')
        else:
            return Enrollment.objects.select_related('student', 'group__course', 'group__teacher')

# Assignment Views
class AssignmentListView(LoginRequiredMixin, ListView):
    model = Assignment
    template_name = 'courses/assignment_list.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def get_queryset(self):
        if hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student':
            student = Student.objects.get(user=self.request.user)
            return Assignment.objects.filter(
                group__enrollments__student=student,
                status='active'
            ).select_related('group__course')
        elif hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'teacher':
            teacher = Teacher.objects.get(user=self.request.user)
            return Assignment.objects.filter(
                group__teacher=teacher
            ).select_related('group__course')
        else:
            return Assignment.objects.select_related('group__course', 'group__teacher')

class AssignmentDetailView(LoginRequiredMixin, DetailView):
    model = Assignment
    template_name = 'courses/assignment_detail.html'
    context_object_name = 'assignment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assignment = self.get_object()
        
        if hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student':
            try:
                student = Student.objects.get(user=self.request.user)
                context['submission'] = Submission.objects.filter(
                    assignment=assignment, student=student
                ).first()
                context['can_submit'] = timezone.now() <= assignment.due_date
            except Student.DoesNotExist:
                pass
        
        context['submissions'] = assignment.submissions.select_related('student')
        return context

class AssignmentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Assignment
    form_class = AssignmentForm
    template_name = 'courses/assignment_form.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'teacher'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('courses:assignment_detail', kwargs={'pk': self.object.pk})

# Submission Views
class SubmissionCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Submission
    form_class = SubmissionForm
    template_name = 'courses/submission_form.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assignment'] = get_object_or_404(Assignment, pk=self.kwargs['pk'])
        return context
    
    def form_valid(self, form):
        assignment = get_object_or_404(Assignment, pk=self.kwargs['pk'])
        student = Student.objects.get(user=self.request.user)
        
        # Check if already submitted
        if Submission.objects.filter(assignment=assignment, student=student).exists():
            messages.error(self.request, 'Bu ödevi zaten teslim ettiniz.')
            return redirect('courses:assignment_detail', pk=assignment.pk)
        
        # Check deadline
        if timezone.now() > assignment.due_date:
            messages.error(self.request, 'Ödev teslim süresi dolmuş.')
            return redirect('courses:assignment_detail', pk=assignment.pk)
        
        submission = form.save(commit=False)
        submission.assignment = assignment
        submission.student = student
        submission.save()
        
        messages.success(self.request, 'Ödev başarıyla teslim edildi.')
        return redirect('courses:assignment_detail', pk=assignment.pk)

# Announcement Views
class AnnouncementListView(LoginRequiredMixin, ListView):
    model = Announcement
    template_name = 'courses/announcement_list.html'
    context_object_name = 'announcements'
    paginate_by = 20
    
    def get_queryset(self):
        if hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student':
            student = Student.objects.get(user=self.request.user)
            return Announcement.objects.filter(
                group__enrollments__student=student,
                status='active'
            ).select_related('group__course', 'teacher').order_by('-create_date')
        else:
            return Announcement.objects.filter(status='active').select_related('group__course', 'teacher').order_by('-create_date')

class AnnouncementDetailView(LoginRequiredMixin, DetailView):
    model = Announcement
    template_name = 'courses/announcement_detail.html'
    context_object_name = 'announcement'

class AnnouncementCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Announcement
    form_class = AnnouncementForm
    template_name = 'courses/announcement_form.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'teacher'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        teacher = Teacher.objects.get(user=self.request.user)
        announcement = form.save(commit=False)
        announcement.teacher = teacher
        announcement.save()
        
        messages.success(self.request, 'Duyuru başarıyla oluşturuldu.')
        return redirect('courses:announcement_detail', pk=announcement.pk)

class AnnouncementUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Announcement
    form_class = AnnouncementForm
    template_name = 'courses/announcement_form.html'
    
    def test_func(self):
        announcement = self.get_object()
        return (hasattr(self.request.user, 'userprofile') and 
                self.request.user.userprofile.user_type == 'teacher' and
                announcement.teacher.user == self.request.user) or self.request.user.is_staff
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, 'Duyuru başarıyla güncellendi.')
        return redirect('courses:announcement_detail', pk=self.object.pk)

class AnnouncementDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Announcement
    template_name = 'courses/announcement_delete.html'
    success_url = reverse_lazy('courses:announcement_list')
    
    def test_func(self):
        announcement = self.get_object()
        return (hasattr(self.request.user, 'userprofile') and 
                self.request.user.userprofile.user_type == 'teacher' and
                announcement.teacher.user == self.request.user) or self.request.user.is_staff
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Duyuru başarıyla silindi.')
        return super().delete(request, *args, **kwargs)

# Grade Management Views
class GradeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Enrollment
    form_class = GradeForm
    template_name = 'courses/grade_form.html'
    
    def test_func(self):
        enrollment = self.get_object()
        return (hasattr(self.request.user, 'userprofile') and 
                self.request.user.userprofile.user_type == 'teacher' and
                enrollment.group.teacher.user == self.request.user)
    
    def get_success_url(self):
        return reverse_lazy('courses:group_detail', kwargs={'pk': self.object.group.pk})
    
    def form_valid(self, form):
        messages.success(self.request, f'{self.object.student.full_name} için notlar başarıyla güncellendi.')
        return super().form_valid(form)

# Report Views
class StudentReportView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'courses/student_report.html'
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin', 'teacher']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enrollments'] = Enrollment.objects.select_related('student', 'group__course').order_by('student__first_name')
        return context
    
    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            return self.export_excel()
        elif request.GET.get('export') == 'pdf':
            return self.export_pdf()
        return super().get(request, *args, **kwargs)
    
    def export_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Öğrenci Raporu"
        
        headers = ['Öğrenci', 'Ders', 'Öğretmen', 'Not', 'Devam', 'Durum']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        enrollments = Enrollment.objects.select_related('student', 'group__course', 'group__teacher')
        for row, enrollment in enumerate(enrollments, 2):
            ws.cell(row=row, column=1, value=enrollment.student.full_name)
            ws.cell(row=row, column=2, value=enrollment.group.course.name)
            ws.cell(row=row, column=3, value=enrollment.group.teacher.full_name)
            ws.cell(row=row, column=4, value=enrollment.grade)
            ws.cell(row=row, column=5, value=f"{enrollment.attendance}%")
            ws.cell(row=row, column=6, value=enrollment.get_status_display())
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ogrenci_raporu.xlsx'
        wb.save(response)
        return response

class CourseReportView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'courses/course_report.html'
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin', 'teacher']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['courses'] = Course.objects.annotate(
            enrollment_count=Count('groups__enrollments')
        ).order_by('code')
        return context

# AJAX Views
@login_required
@csrf_exempt
def update_grade_ajax(request):
    """AJAX endpoint for inline grade updates"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'})
    
    # Check if user is teacher
    if not (hasattr(request.user, 'userprofile') and request.user.userprofile.user_type == 'teacher'):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        grade_type = data.get('grade_type')
        score = data.get('score')
        
        # Validate input
        if not all([enrollment_id, grade_type, score is not None]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'})
        
        try:
            score = float(score)
            if score < 0 or score > 100:
                return JsonResponse({'success': False, 'error': 'Score must be between 0 and 100'})
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid score format'})
        
        # Get enrollment
        try:
            enrollment = Enrollment.objects.get(id=enrollment_id)
        except Enrollment.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Enrollment not found'})
        
        # Check if teacher owns this course
        teacher = Teacher.objects.get(user=request.user)
        if enrollment.group.teacher != teacher:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        
        # Update or create note
        from notes.models import Note
        
        # First update the enrollment model
        if grade_type == 'vize':
            enrollment.midterm_grade = score
        elif grade_type == 'final':
            enrollment.final_grade = score
        elif grade_type == 'but':
            enrollment.makeup_grade = score
        else:
            return JsonResponse({'success': False, 'error': 'Invalid grade type'})
        
        enrollment.save()  # This will auto-calculate letter grade
        
        # Also update/create the Note model for consistency
        note, created = Note.objects.update_or_create(
            student=enrollment.student.user,
            course=enrollment.group.course,
            exam_type=grade_type,
            defaults={
                'score': int(score),
                'teacher': request.user
            }
        )
        
        return JsonResponse({
            'success': True,
            'letter_grade': enrollment.grade,
            'message': 'Grade updated successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Teacher.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Teacher profile not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})