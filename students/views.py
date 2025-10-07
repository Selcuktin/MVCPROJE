from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.http import HttpResponse
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from .models import Student
from courses.models import Enrollment, Assignment, Submission, Announcement
from .forms import StudentForm

class StudentListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'students/list.html'
    context_object_name = 'students'
    paginate_by = 20
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['teacher', 'admin']
    
    def get_queryset(self):
        queryset = Student.objects.all()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(tc_no__icontains=search) |
                Q(email__icontains=search)
            )
        
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('first_name', 'last_name')

class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = 'students/detail.html'
    context_object_name = 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        context['enrollments'] = student.enrollments.select_related('group__course', 'group__teacher')
        context['submissions'] = student.submissions.select_related('assignment__group__course')
        return context

class StudentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'students/form.html'
    success_url = reverse_lazy('students:list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin']
    
    def form_valid(self, form):
        messages.success(self.request, 'Öğrenci başarıyla oluşturuldu.')
        return super().form_valid(form)

class StudentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'students/form.html'
    success_url = reverse_lazy('students:list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['teacher', 'admin']
    
    def form_valid(self, form):
        messages.success(self.request, 'Öğrenci bilgileri başarıyla güncellendi.')
        return super().form_valid(form)

class StudentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Student
    template_name = 'students/delete.html'
    success_url = reverse_lazy('students:list')
    
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['admin']
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Öğrenci başarıyla silindi.')
        return super().delete(request, *args, **kwargs)

class StudentDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'students/dashboard.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student'
    
    def handle_no_permission(self):
        messages.error(self.request, 'Öğrenci paneline sadece öğrenciler erişebilir.')
        return redirect('home')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            student = Student.objects.get(user=self.request.user)
            context['student'] = student
            context['enrollments'] = student.enrollments.select_related('group__course', 'group__teacher')
            context['recent_assignments'] = Assignment.objects.filter(
                group__enrollments__student=student,
                status='active'
            ).order_by('-create_date')[:5]
            context['recent_announcements'] = Announcement.objects.filter(
                group__enrollments__student=student,
                status='active'
            ).select_related('group__course').order_by('-create_date')[:5]
            context['pending_submissions'] = Assignment.objects.filter(
                group__enrollments__student=student,
                status='active'
            ).exclude(
                submissions__student=student
            ).count()
        except Student.DoesNotExist:
            context['student'] = None
        return context

class StudentCoursesView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'students/courses.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            student = Student.objects.get(user=self.request.user)
            context['enrollments'] = student.enrollments.select_related('group__course', 'group__teacher')
        except Student.DoesNotExist:
            context['enrollments'] = []
        return context

class StudentAssignmentsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'students/assignments.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            student = Student.objects.get(user=self.request.user)
            assignments = Assignment.objects.filter(
                group__enrollments__student=student
            ).select_related('group__course').order_by('-create_date')
            
            submissions = {
                sub.assignment_id: sub for sub in 
                Submission.objects.filter(student=student)
            }
            
            # Add submission data to each assignment
            assignments_with_submissions = []
            for assignment in assignments:
                assignment.submission = submissions.get(assignment.id)
                assignments_with_submissions.append(assignment)
            
            context['assignments'] = assignments_with_submissions
            context['submissions'] = submissions
        except Student.DoesNotExist:
            context['assignments'] = []
            context['submissions'] = {}
        return context

class StudentGradesView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'students/grades.html'
    
    def test_func(self):
        return hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type == 'student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            student = Student.objects.get(user=self.request.user)
            context['enrollments'] = student.enrollments.select_related('group__course', 'group__teacher')
            context['submissions'] = student.submissions.select_related('assignment__group__course')
        except Student.DoesNotExist:
            context['enrollments'] = []
            context['submissions'] = []
        return context

class StudentExportView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    def test_func(self):
        return self.request.user.is_staff or hasattr(self.request.user, 'userprofile') and self.request.user.userprofile.user_type in ['teacher', 'admin']
    
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'excel')
        
        if format_type == 'pdf':
            return self.export_pdf()
        else:
            return self.export_excel()
    
    def export_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Öğrenci Listesi"
        
        # Headers
        headers = ['TC No', 'Ad', 'Soyad', 'Email', 'Telefon', 'Durum', 'Kayıt Tarihi']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        students = Student.objects.all()
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student.tc_no)
            ws.cell(row=row, column=2, value=student.first_name)
            ws.cell(row=row, column=3, value=student.last_name)
            ws.cell(row=row, column=4, value=student.email)
            ws.cell(row=row, column=5, value=student.phone)
            ws.cell(row=row, column=6, value=student.get_status_display())
            ws.cell(row=row, column=7, value=student.registration_date.strftime('%d.%m.%Y'))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ogrenci_listesi.xlsx'
        wb.save(response)
        return response
    
    def export_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=ogrenci_listesi.pdf'
        
        p = canvas.Canvas(response, pagesize=letter)
        p.drawString(100, 750, "Öğrenci Listesi")
        
        students = Student.objects.all()[:20]  # Limit for demo
        y = 700
        for student in students:
            p.drawString(100, y, f"{student.tc_no} - {student.full_name}")
            y -= 20
        
        p.showPage()
        p.save()
        return response